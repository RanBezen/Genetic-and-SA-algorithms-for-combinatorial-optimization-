import argparse
import glob
from utils_comb import bruteForce_CommExpand, greedy_backward_CommExpand, greedy_forward_CommExpand, joker_CommExpand,Genetic_Algorithms_CommExpand,bruteForce_graphCond
import os
from openpyxl import Workbook, load_workbook

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--mode', type=str, default='',
                        help='What operation? Can choose name or parse')
    parser.add_argument('--infile', type=str, default='inputComm.txt',
                        help='insert path')
    parser.add_argument('--algorithm', type=str, default='',
                        help='algorithm name')
    parser.add_argument('--timeout', type=float, default=60,
                        help='algorithm name')
    args = parser.parse_args()
    mode(args)

def mode(args):
    if args.mode == 'name':
        print('Ran')
    elif args.mode == 'parse':
        parse(args.infile)
    elif args.mode == 'solve':
        if args.algorithm == 'brute_force':
            brute_force(args.infile)
        if args.algorithm == 'greedy':
            greedy(args.infile)
        if args.algorithm == 'joker':
            joker(args.infile)
    elif args.mode == 'compete':
        compete(args.infile, args.timeout)


def parse(path):
    edges_list, vertices_A, phi = read_txt(path)
    vertices_list_flat = [val for sublist in edges_list for val in sublist]
    vertices_list = set(vertices_list_flat)
    edges_A = []
    for edge in edges_list:
        # check if list1 contains all elements in list2
        result = all(elem in set(vertices_A) for elem in set(edge))
        # if set(vertices_A) >= set(edge):
        if result:
            edges_A.append(edge)
    print(len(vertices_list))
    print(len(edges_list))
    print(len(vertices_A))
    print(len(edges_A))
    print(phi)

def read_txt(path):
    """
    read txt file to edges_list, vertices_A and phi
    :param path:
    :return: edges_list, vertices_A, phi
    """
    file = open(path, 'r')
    rows = file.read().splitlines()
    edges_list = []
    flag_A = False
    flag_phi = False
    vertices_A = []
    for row in rows:
        edge = row.split(' ')
        if edge[0] == 'A':
            flag_A = True
            edge = edge[1:]
        if edge[0] == 'phi':
            flag_phi = True
        if not flag_A and not flag_phi:
            temp = []
            for ver in edge:
                temp.append(int(ver))
            edges_list.append(temp)
        if flag_A and not flag_phi:
            for ver in edge:
                if ver!='':
                    vertices_A.append(int(ver))
        if flag_phi:
            phi = float(edge[1])
    return edges_list, vertices_A, phi

def brute_force(path):
    edges, vertices_A, phi = read_txt(path)
    maxA_vertices, maxA_phi, edge_to_cut = bruteForce_CommExpand(edges, vertices_A,phi)
    maxA_vertices = set(maxA_vertices)
    str_A = ' '.join(str(e) for e in maxA_vertices)
    #print(str_A, '|',len(maxA_vertices), '|',maxA_phi)
    print(str_A)
    return maxA_vertices


def brute_force_under(path,under):
    edges, vertices_A, phi = read_txt(path)
    vertices = [val for sublist in edges for val in sublist]  # get all vertices
    vertices = set(vertices)
    if len(vertices) < under:
        maxA_vertices, maxA_phi, edge_to_cut = bruteForce_CommExpand(edges, vertices_A,phi)
        maxA_vertices = set(maxA_vertices)
        str_A = ' '.join(str(e) for e in maxA_vertices)
        #print(str_A, '|',len(maxA_vertices), '|',maxA_phi)
        #print('brute force:',str_A, '|',len(maxA_vertices))
    else:
        maxA_vertices = [0,0,0,0]
    return maxA_vertices


def greedy(path):
    edges, vertices_A, phi = read_txt(path)
    maxA_vertices, maxA_phi, edge_to_cut = greedy_backward_CommExpand(edges, vertices_A,phi)
    str_A = ' '.join(str(e) for e in maxA_vertices)
    #print('greedy backward:',str_A, '|',len(maxA_vertices))
    print(str_A)

def joker(path):
    edges, vertices_A, phi = read_txt(path)
    maxA_vertices, maxA_phi, edge_to_cut = joker_CommExpand(edges, vertices_A, phi)
    str_A = ' '.join(str(e) for e in maxA_vertices)
    print(str_A)


def compete(path,timeout):
    edges, vertices_A, phi = read_txt(path)
    maxA_vertices, maxA_phi = Genetic_Algorithms_CommExpand(edges, vertices_A, phi,timeout)
    str_A = ' '.join(str(e) for e in maxA_vertices)
    print(str_A)
    #print(str_A, '|',len(maxA_vertices), '|',maxA_phi)
    return maxA_vertices, edges, phi


def run_all(path):
    greedy(path); print('greedy')
    joker(path); print('joker')
    brute_force(path); print('brute_force')
    compete(path, timeout=10); print('Genetic_Algorithms')

def test_on_batch():
    folders = ['26','40-50','80-100']
    # folders = ['26']
    for folder in folders:
        if folder == '26':
            timeouts = [1, 5, 10, 30, 60, 120]
        else:
            timeouts = [15, 30, 60, 5 * 60, 10 * 60, 20 * 60]

        for infile in glob.glob1('D:/PycharmProjects/optComb/test/comm/' + folder, 'inputComm*.txt'):
            path = 'test/comm/' + folder + '/' + infile
            print(infile, '#####################')
            temp_data = []
            temp_data2 = []
            temp_data3 = []
            temp_data.append(infile.split('.')[0])
            temp_data2.append(infile.split('.')[0])
            temp_data3.append(infile.split('.')[0])

            for time in timeouts:
                A, edges, phi_original = compete(path, timeout=time)
                if folder == '26':
                    Aphi_pred, _ = bruteForce_graphCond(edges, A)
                    temp_data3.append(Aphi_pred)
                str_A = ' '.join(str(e) for e in A)
                temp_data.append(len(A))
                temp_data2.append(str_A)
                print(temp_data, temp_data2, temp_data3)
            if folder == '26':
                A_brute = brute_force(path)
                str_A = ' '.join(str(e) for e in A_brute)
                temp_data.append(len(A_brute))
            filepath = 'comm_compete_results_rangeOf_' + folder + '.xlsx'
            if not os.path.exists(filepath):
                wb = Workbook()
                sheet = wb.active
                headers = list(map(str, timeouts))
                headers.insert(0, 'name')
                sheet.append(headers)
                source = wb['Sheet']
                target1 = wb.copy_worksheet(source)
                target2 = wb.copy_worksheet(source)
                source.title = 'A size'
                target1.title = 'A ver'
                target2.title = 'A phi brute'
                wb.save(filepath)
            wb = load_workbook(filepath)
            source1 = wb['A size']
            source2 = wb['A ver']
            source3 = wb['A phi brute']
            source1.append(temp_data)
            source2.append(temp_data2)
            source3.append(temp_data3)
            wb.save(filepath)


if __name__ == '__main__':
    main()
