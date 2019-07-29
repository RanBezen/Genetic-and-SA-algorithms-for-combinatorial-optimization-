import argparse
from utils_comb import bruteForce_graphCond, greedy_graphCond, joker_graphCond, simulated_annealing_graphCond, lower_bound_cheeger,greedy_graphCond_smart
import glob
import os
from openpyxl import Workbook, load_workbook

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--mode', type=str, default='',
                        help='What operation? Can choose name or parse')
    parser.add_argument('--infile', type=str, default='inputGraph.txt',
                        help='insert path')
    parser.add_argument('--algorithm', type=str, default='',
                        help='algorithm name')
    parser.add_argument('--timeout', type=float, default=60,
                        help='algorithm name')
    args = parser.parse_args()
    mode(args)

def mode(args):
    if args.mode == 'name':
        print('Ran')
    elif args.mode == 'parse':
        parse(args.infile)
    elif args.mode == 'solve':
        if args.algorithm == 'brute_force':
            brute_force(args.infile)
        if args.algorithm == 'greedy':
            greedy(args.infile)
        if args.algorithm == 'joker':
            joker(args.infile)
    elif args.mode == 'compete':
        compete(args.infile, args.timeout)

def parse(path):

    edges_list = read_txt(path)
    vertices_list_flat = [val for sublist in edges_list for val in sublist]
    vertices_list = set(vertices_list_flat)
    print(len(vertices_list))
    print(len(edges_list))

def read_txt(path):
    file = open(path, 'r')
    rows = file.read().splitlines()
    edges_list = []
    for row in rows:
        edge = row.split(' ')
        temp = []
        for ver in edge:
            temp.append(int(ver))
        edges_list.append(temp)
    return edges_list

def brute_force(path):
    edges = read_txt(path)
    phi,edge_cut = bruteForce_graphCond(edges)
    print(str(phi))
    return str(phi)

def greedy(path):
    edges = read_txt(path)
    phi,_,edge_cut = greedy_graphCond_smart(edges,random_ver=True)
    print(str(phi))

def joker(path):
    edges = read_txt(path)
    phi, _,edge_cut = joker_graphCond(edges)
    print(str(phi))

def compete(path,timeout):
    edges = read_txt(path)
    phi, cutted_edges_min = simulated_annealing_graphCond(edges, timeout)
    #print('compete:',phi)
    print(str(phi))
    return phi

def run_all(path):
    greedy(path)
    joker(path)

def lower_bound(path):
    edges = read_txt(path)
    lower,upper = lower_bound_cheeger(edges)
    print(str(lower),str(upper))

def test_on_batch():
    folders = ['26','40-50', '80-100','400-500']
    # folders = ['26']
    for folder in folders:
        if folder == '26':
            timeouts = [1, 5, 10, 30, 60]
        else:
            timeouts = [15, 30, 60, 5 * 60, 10 * 60, 20 * 60]
        for infile in glob.glob1('D:/PycharmProjects/optComb/test/graph_con/' + folder, 'inputGraph_*.txt'):
            path = 'test/graph_con/' + folder + '/' + infile
            print(infile, '#####################')
            temp_data = []
            temp_data.append(infile.split('.')[0])
            for time in timeouts:
                phi = compete(path, timeout=time)
                # temp_data = [infile.split('.')[0], phi, time]
                temp_data.append(phi)
                print(temp_data)

            if folder == '26':
                phi_brute = brute_force(path)
                temp_data.append(phi_brute)
                print(temp_data)

            filepath = 'cond_compete_results_rangeOf_' + folder + '.xlsx'
            if not os.path.exists(filepath):
                wb = Workbook()
                sheet = wb.active
                headers = list(map(str, timeouts))
                headers.insert(0, 'name')
                sheet.append(headers)
                wb.save(filepath)
            wb = load_workbook(filepath)
            sheet = wb.active
            sheet.append(temp_data)
            wb.save(filepath)
if __name__ == '__main__':
    main()
