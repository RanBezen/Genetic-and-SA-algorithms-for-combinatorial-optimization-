import numpy as np
import random
import operator as op
from functools import reduce
from itertools import combinations
import argparse
from utils_comb import bruteForce_graphCond, collect_edges_A
import os
from openpyxl import Workbook, load_workbook
import copy

def touched_ver(edges,vertices_A_original):
    vertices = [val for sublist in edges for val in sublist]  # get all vertices
    vertices = set(vertices)
    verticesB = [item for item in vertices if item not in vertices_A_original]  # get vertices of A_ B
    touch_ver = []
    for w in range(len(edges)):
        edge = edges[w]
        CutFlag = any(elem in set(vertices_A_original) for elem in set(edge))
        GroupAflag = all(elem in set(vertices_A_original) for elem in set(edge))
        GroupBflag = all(elem in set(verticesB) for elem in set(edge))
        if CutFlag and not GroupAflag and not GroupBflag:  # check if the edge is in cut
            touch_ver.append(edge)
    touch_ver = set([val for sublist in touch_ver for val in sublist])
    touch_ver = [item for item in touch_ver if item not in vertices_A_original] #get vertices of A_ B
    return touch_ver

def graph_generator(n, p, A_number,gen_community=True):
    """
    :param n: number of vertices
    :param p: number of edges
    :param A_number: number of vertices in A
    :return: edges, A, phi
    """

    from_n = 1
    to_n = n+1
    vertices = random.sample(range(from_n, to_n), n)
    edges = get_random_edges(vertices, p)
    if gen_community:
        A, Aphi = get_random_A_Phi(edges,A_number)
    else:
        A = []
        Aphi = 0
    return edges, A, Aphi

def get_random_A_Phi(edges,A_number):

    edge1 = random.sample(edges,1)[0]
    A = edge1
    while len(A)<A_number:
        neighburs = touched_ver(edges, A)
        if len(neighburs)>1:
            neighburs = random.sample(neighburs,max(1,int(len(neighburs)/5)))
        A.extend(neighburs)
    A_edges = collect_edges_A(edges, A)
    phi, _ = bruteForce_graphCond(A_edges)
    phi = phi*0.8
    phi = round(phi,2)
    return A, phi

def get_random_edges(vertices_org, p):
    # Generate all possible non-repeating pairs
    #vertices = copy.deepcopy(vertices_org)
    vertices = vertices_org
    pairs_gen = list(combinations(vertices, 2))
    random.shuffle(pairs_gen)
    edges = []

    end_i = len(pairs_gen)
    pairs = []
    i_vec = range(1, end_i)
    for pair1 in pairs_gen:
        pairs.append(list(pair1))

    edges.append(pairs[0])
    pairs.remove(pairs[0])

    v = [val for sublist in edges for val in sublist]
    set_v = set(v)
    l=len(vertices)
    len_edges = len(edges)
    set_vertices = set(vertices)
    while not all(elem in set_v for elem in set_vertices) or len_edges < p:
        touch_ver = []
        for w in range(len(pairs)):
            edge = pairs[w]
            CutFlag = any(elem in set(v) for elem in set(edge))
            #GroupAflag = all(elem in set(v) for elem in set(edge))
            GroupAflag = False
            for a in edges:
                 if a == edge:
                    GroupAflag = True
                    break
            if CutFlag and not GroupAflag:  # check if the edge is in cut
                touch_ver.append(edge)

        i = random.sample(range(len(touch_ver)), 1)[0]
        pair = touch_ver[i]

        edges.append(pair)
        pairs.remove(pair)
        len_edges = len(edges)
        v = [val for sublist in edges for val in sublist]
        set_v = set(v)
        # i+=1
    return edges

def write_txt(name,edges, A=[], phi=0):
    with open(name+'.txt', 'w') as f:
        for item in edges:
            item_str = str(item[0]) + ' ' + str(item[1])
            f.write("%s\n" % item_str)
        if A:
            A_str = ''
            for i in A:
                A_str += str(i) + ' '
            A_str = A_str[0:len(A_str) - 1]
            f.write("A %s\n" % A_str)
            f.write("phi %s\n" % str(phi))

def ncr(n, r):
    r = min(r, n-r)
    numer = reduce(op.mul, range(n, n-r, -1), 1)
    denom = reduce(op.mul, range(1, r+1), 1)
    return round(numer / denom)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--number_of_txt', type=int, default=5,
                        help='Please insert...')
    parser.add_argument('--upper_limit_vertices', type=int, default=12,
                        help='...')
    parser.add_argument('--lower_limit_vertices', type=int, default=4,
                        help='...')
    args = parser.parse_args()

    # number_of_txt = args.number_of_txt
    # upper_limit_vertices = args.upper_limit_vertices
    # lower_limit_vertices = args.lower_limit_vertices

    filepath = 'test data.xlsx'
    number_of_txt=10
    upper_limit_vertices=500
    lower_limit_vertices=400



    # for i in range(1,number_of_txt+1):
    #
    #     n=np.random.randint(lower_limit_vertices, upper_limit_vertices)
    #     #p=np.random.randint(round(n/2)+1,ncr(n, 2))
    #     p=ncr(n, 2)
    #     A_number=np.random.randint(1,round(n/2)+1)
    #     edges, A, phi = graph_generator(n, p, A_number)
    #     inputComm_name = 'inputComm_2'+str(i)
    #     inputGraph_name = 'inputGraph_2'+str(i)
    #     write_txt(inputComm_name, edges, A, phi)
    #     write_txt(inputGraph_name, edges)

    k = 0
    while k < number_of_txt:
        n = np.random.randint(lower_limit_vertices, upper_limit_vertices)
        #p=np.random.randint(round(n)-1,ncr(n, 2))
        p = np.random.randint(round(n)-1,round(n*3))
        #A_number = np.random.randint(4, round(n / 2) + 1)
        A_number = np.random.randint(4, min(round(n / 2) + 1, 11))
        print('try  | n:', n, '| p:', p, '| A_number:', A_number)
        edges, A, Aphi = graph_generator(n, p, A_number, gen_community=False)

        #phi_graph, _ = bruteForce_graphCond(edges)
        #if phi_graph > 0 and phi_graph < Aphi:
        if True:
            vertices = [val for sublist in edges for val in sublist]  # get all vertices
            vertices = list(set(vertices))
            inputComm_name = 'inputComm_large3' + str(k)
            inputGraph_name = 'inputGraph_large3' + str(k)
            #write_txt(inputComm_name, edges, A, Aphi)
            write_txt(inputGraph_name, edges)
            k+=1

            #temp_data = [inputComm_name, len(edges), len(vertices),len(vertices)/len(edges) , phi_graph,len(A), Aphi]
            temp_data = [inputComm_name, len(edges), len(vertices),len(vertices)/len(edges), len(A), Aphi]

            print(temp_data)
            if not os.path.exists(filepath):
                wb = Workbook()
                sheet = wb.active
                #sheet.append(['name', 'edges', 'vertices','retio' ,'phi_graph' ,'A vertices', 'Aphi'])
                sheet.append(['name', 'edges', 'vertices','retio','A vertices', 'Aphi'])

                wb.save(filepath)
            wb = load_workbook(filepath)
            sheet = wb.active
            sheet.append(temp_data)
            wb.save(filepath)
        #else:
            #print('phi_graph > Aphi:', phi_graph > Aphi, 'phi_graph > 0:', phi_graph > 0)